import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkcalendar import Calendar
import datetime
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── Palette ──────────────────────────────────────────────────────────────────
BG        = "#1e1e2e"
PANEL     = "#2a2a3d"
ACCENT    = "#7c6af7"
ACCENT_HV = "#9d8fff"
SUCCESS   = "#50fa7b"
ERROR_CLR = "#ff5555"
FG        = "#cdd6f4"
FG_DIM    = "#888aad"
ENTRY_BG  = "#313244"
BTN_FG    = "#ffffff"

FONT_TITLE  = ("Segoe UI", 13, "bold")
FONT_LABEL  = ("Segoe UI", 10)
FONT_SMALL  = ("Segoe UI", 9)
FONT_STATUS = ("Segoe UI", 9, "italic")
FONT_BTN    = ("Segoe UI", 10, "bold")


def styled_button(parent, text, command, bg=ACCENT, width=18):
    btn = tk.Button(
        parent, text=text, command=command,
        bg=bg, fg=BTN_FG, activebackground=ACCENT_HV, activeforeground=BTN_FG,
        font=FONT_BTN, relief="flat", cursor="hand2",
        padx=10, pady=6, width=width, bd=0
    )
    btn.bind("<Enter>", lambda e: btn.config(bg=ACCENT_HV))
    btn.bind("<Leave>", lambda e: btn.config(bg=bg))
    return btn


def _write_xlsx(path, rows):
    """Write schedule rows to an Excel file with styled headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Schedule"

    # Header styling
    header_fill = PatternFill("solid", fgColor="313244")
    header_font = Font(name="Segoe UI", bold=True, color="7C6AF7", size=11)
    cell_font   = Font(name="Segoe UI", size=10)

    headers = ["datetime", "file_path"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r, (dt, fp) in enumerate(rows, start=2):
        ws.cell(row=r, column=1, value=dt).font = cell_font
        ws.cell(row=r, column=2, value=fp).font = cell_font

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60

    wb.save(path)


def _read_xlsx(path):
    """Return list of (datetime_str, file_path) tuples from an Excel file."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        dt  = row[0] if row[0] is not None else ""
        fp  = row[1] if len(row) > 1 and row[1] is not None else ""
        if dt or fp:
            rows.append((str(dt), str(fp)))
    return rows


class XLSXManagerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Fair Announcement Scheduler  —  Excel Builder")
        self.root.configure(bg=BG)
        self.root.resizable(True, True)
        self.root.minsize(800, 580)

        self.filename = ""
        self.selected_time = None
        self._tooltip = None
        self._schedule_rows = []   # list of (datetime_str, file_path)

        self._build_ui()
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

    # ── Main layout ───────────────────────────────────────────────────────────

    def _build_ui(self):
        outer = tk.Frame(self.root, bg=BG, padx=18, pady=18)
        outer.grid(sticky="nsew")
        outer.columnconfigure(0, weight=1)
        outer.rowconfigure(2, weight=1)

        self._build_header(outer)
        self._build_form(outer)
        self._build_table(outer)
        self._build_statusbar(outer)

    def _build_header(self, parent):
        hdr = tk.Frame(parent, bg=BG)
        hdr.grid(row=0, column=0, sticky="ew", pady=(0, 14))
        hdr.columnconfigure(1, weight=1)

        tk.Label(hdr, text="📢  Fair Announcement Scheduler  •  Excel Builder",
                 font=FONT_TITLE, bg=BG, fg=FG).grid(row=0, column=0, sticky="w")

        file_row = tk.Frame(hdr, bg=BG)
        file_row.grid(row=0, column=1, sticky="e")

        self.file_label = tk.Label(file_row, text="No schedule file selected",
                                   font=FONT_SMALL, bg=BG, fg=FG_DIM,
                                   anchor="e", width=42)
        self.file_label.grid(row=0, column=0, padx=(0, 8))

        styled_button(file_row, "📂  Open / New XLSX", self.select_file,
                      width=20).grid(row=0, column=1)

    def _build_form(self, parent):
        card = tk.Frame(parent, bg=PANEL, padx=16, pady=14,
                        highlightbackground=ACCENT, highlightthickness=1)
        card.grid(row=1, column=0, sticky="ew", pady=(0, 14))
        card.columnconfigure(1, weight=1)
        card.columnconfigure(3, weight=1)

        tk.Label(card, text="ADD ANNOUNCEMENT", font=("Segoe UI", 9, "bold"),
                 bg=PANEL, fg=ACCENT).grid(row=0, column=0, columnspan=4,
                                            sticky="w", pady=(0, 10))

        tk.Label(card, text="Date & Time:", font=FONT_LABEL,
                 bg=PANEL, fg=FG).grid(row=1, column=0, sticky="w", padx=(0, 8))

        self.datetime_entry = tk.Entry(
            card, font=FONT_LABEL, bg=ENTRY_BG, fg=FG,
            insertbackground=FG, relief="flat", width=22
        )
        self.datetime_entry.grid(row=1, column=1, sticky="ew", ipady=5)

        styled_button(card, "📅  Pick Date & Time",
                      self.pick_datetime, width=20).grid(row=1, column=2, padx=(10, 0))

        tk.Label(card, text="Audio File:", font=FONT_LABEL,
                 bg=PANEL, fg=FG).grid(row=2, column=0, sticky="w",
                                        padx=(0, 8), pady=(10, 0))

        self.filepath_entry = tk.Entry(
            card, font=FONT_LABEL, bg=ENTRY_BG, fg=FG,
            insertbackground=FG, relief="flat", width=22
        )
        self.filepath_entry.grid(row=2, column=1, sticky="ew", ipady=5, pady=(10, 0))

        styled_button(card, "🔍  Browse",
                      self.browse_file, width=20).grid(row=2, column=2,
                                                        padx=(10, 0), pady=(10, 0))

        btn_row = tk.Frame(card, bg=PANEL)
        btn_row.grid(row=3, column=0, columnspan=3, pady=(14, 0))

        styled_button(btn_row, "＋  Add to Schedule",
                      self.add_entry, width=22).pack(side="left", padx=(0, 10))
        styled_button(btn_row, "💾  Save Excel File",
                      self.save_file, bg="#27ae60", width=20).pack(side="left")

    def _build_table(self, parent):
        table_frame = tk.Frame(parent, bg=BG)
        table_frame.grid(row=2, column=0, sticky="nsew")
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(1, weight=1)

        hdr = tk.Frame(table_frame, bg=BG)
        hdr.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        hdr.columnconfigure(0, weight=1)

        tk.Label(hdr, text="SCHEDULE", font=("Segoe UI", 9, "bold"),
                 bg=BG, fg=ACCENT).grid(row=0, column=0, sticky="w")

        styled_button(hdr, "🗑  Delete Selected",
                      self.delete_selected,
                      bg="#c0392b", width=18).grid(row=0, column=1, sticky="e")

        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Custom.Treeview",
                        background=PANEL, foreground=FG,
                        fieldbackground=PANEL, rowheight=28,
                        font=FONT_LABEL, borderwidth=0)
        style.configure("Custom.Treeview.Heading",
                        background=ENTRY_BG, foreground=ACCENT,
                        font=("Segoe UI", 10, "bold"), relief="flat")
        style.map("Custom.Treeview",
                  background=[("selected", ACCENT)],
                  foreground=[("selected", BTN_FG)])

        cols = ("datetime", "file_path")
        self.tree = ttk.Treeview(table_frame, columns=cols,
                                  show="headings", style="Custom.Treeview",
                                  selectmode="browse")
        self.tree.heading("datetime",  text="Scheduled Date & Time")
        self.tree.heading("file_path", text="Audio File")
        self.tree.column("datetime",  width=200, anchor="center", stretch=False)
        self.tree.column("file_path", width=400, anchor="w",      stretch=True)

        vsb = ttk.Scrollbar(table_frame, orient="vertical",
                             command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)

        self.tree.grid(row=1, column=0, sticky="nsew")
        vsb.grid(row=1, column=1, sticky="ns")

        self.tree.tag_configure("odd",  background=PANEL)
        self.tree.tag_configure("even", background="#24243a")

        self.tree.bind("<Motion>", self._on_tree_motion)
        self.tree.bind("<Leave>",  self._hide_tooltip)

    def _build_statusbar(self, parent):
        self.status_var = tk.StringVar(value="Ready.")
        bar = tk.Label(parent, textvariable=self.status_var,
                       font=FONT_STATUS, bg=BG, fg=FG_DIM, anchor="w")
        bar.grid(row=3, column=0, sticky="ew", pady=(8, 0))
        self._status_bar = bar

    # ── Status helper ─────────────────────────────────────────────────────────

    def _set_status(self, msg, color=FG_DIM):
        self.status_var.set(msg)
        self._status_bar.config(fg=color)

    # ── File helpers ──────────────────────────────────────────────────────────

    def select_file(self):
        path = filedialog.askopenfilename(
            title="Open or select an Excel schedule file",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not path:
            # Allow creating a new file by asking for a save path
            path = filedialog.asksaveasfilename(
                title="Create new Excel schedule file",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
        if not path:
            return

        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        self.filename = path
        short = os.path.basename(path)
        self.file_label.config(text=short, fg=FG)

        if os.path.exists(path):
            self._reload_table()
            self._set_status(f"Loaded: {path}", SUCCESS)
        else:
            self._schedule_rows = []
            self._set_status(f"New file: {path}  (save when ready)", FG_DIM)

    def browse_file(self):
        path = filedialog.askopenfilename(
            title="Select audio file",
            filetypes=[("Audio files", "*.mp3 *.wav *.ogg *.flac *.m4a"),
                       ("All files", "*.*")]
        )
        if path:
            self.filepath_entry.delete(0, tk.END)
            self.filepath_entry.insert(0, path)

    def save_file(self):
        if not self.filename:
            self._set_status("Please open or create an Excel file first.", ERROR_CLR)
            return
        try:
            _write_xlsx(self.filename, self._schedule_rows)
            self._set_status(f"Saved: {self.filename}", SUCCESS)
        except Exception as e:
            self._set_status(f"Could not save file: {e}", ERROR_CLR)

    # ── Table helpers ─────────────────────────────────────────────────────────

    def _reload_table(self):
        self.tree.delete(*self.tree.get_children())
        self._schedule_rows = []
        if not self.filename or not os.path.exists(self.filename):
            return
        try:
            self._schedule_rows = _read_xlsx(self.filename)
            for i, (dt, fp) in enumerate(self._schedule_rows):
                tag = "odd" if i % 2 else "even"
                self.tree.insert("", "end", values=(dt, fp), tags=(tag,))
        except Exception as e:
            self._set_status(f"Could not read Excel file: {e}", ERROR_CLR)

    def delete_selected(self):
        sel = self.tree.selection()
        if not sel:
            self._set_status("No row selected.", FG_DIM)
            return
        item = sel[0]
        values = self.tree.item(item, "values")
        if not messagebox.askyesno("Confirm Delete",
                                   f"Remove this entry?\n\n"
                                   f"  Date/Time : {values[0]}\n"
                                   f"  File      : {values[1]}"):
            return

        # Remove from internal list and tree
        self._schedule_rows = [
            (dt, fp) for dt, fp in self._schedule_rows
            if not (dt == values[0] and fp == values[1])
        ]
        self.tree.delete(item)
        self._restripe()
        self._set_status("Entry removed. Click 'Save Excel File' to persist.", SUCCESS)

    def _restripe(self):
        for i, item in enumerate(self.tree.get_children()):
            tag = "odd" if i % 2 else "even"
            self.tree.item(item, tags=(tag,))

    # ── Tooltip for file paths ────────────────────────────────────────────────

    def _on_tree_motion(self, event):
        row_id = self.tree.identify_row(event.y)
        col_id = self.tree.identify_column(event.x)
        if row_id and col_id == "#2":
            values = self.tree.item(row_id, "values")
            if values:
                self._show_tooltip(event, values[1])
                return
        self._hide_tooltip()

    def _show_tooltip(self, event, text):
        if self._tooltip:
            self._tooltip.destroy()
        x = event.x_root + 14
        y = event.y_root + 14
        self._tooltip = tw = tk.Toplevel(self.root)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        tk.Label(tw, text=text, font=FONT_SMALL,
                 bg="#333355", fg=FG, relief="solid", bd=1,
                 padx=6, pady=4).pack()

    def _hide_tooltip(self, *_):
        if self._tooltip:
            self._tooltip.destroy()
            self._tooltip = None

    # ── Add entry ─────────────────────────────────────────────────────────────

    def add_entry(self):
        if not self.filename:
            self._set_status("Please open or create an Excel file first.", ERROR_CLR)
            return

        date_time = self.datetime_entry.get().strip()
        file_path = self.filepath_entry.get().strip()

        if not date_time or not file_path:
            self._set_status("Date/time and audio file are required.", ERROR_CLR)
            return

        # Validate date format
        try:
            datetime.datetime.strptime(date_time, "%m/%d/%Y %H:%M")
        except ValueError:
            self._set_status("Date/time must be MM/DD/YYYY HH:MM format.", ERROR_CLR)
            return

        if any(dt == date_time and fp == file_path
               for dt, fp in self._schedule_rows):
            self._set_status("Duplicate entry — not added.", ERROR_CLR)
            return

        self._schedule_rows.append((date_time, file_path))

        count = len(self.tree.get_children())
        tag = "odd" if count % 2 else "even"
        self.tree.insert("", "end", values=(date_time, file_path), tags=(tag,))

        self.datetime_entry.delete(0, tk.END)
        self.filepath_entry.delete(0, tk.END)
        self._set_status(
            f"Added: {date_time}  →  {os.path.basename(file_path)}  "
            f"(click 'Save Excel File' to write to disk)",
            SUCCESS
        )

    # ── Date / Time picker ────────────────────────────────────────────────────

    def pick_datetime(self):
        self.selected_time = None

        win = tk.Toplevel(self.root)
        win.title("Pick Date & Time")
        win.configure(bg=BG)
        win.resizable(False, False)
        win.grab_set()
        self.cal_window = win

        tk.Label(win, text="Select Date", font=FONT_LABEL,
                 bg=BG, fg=FG).pack(pady=(14, 4))

        self.cal = Calendar(
            win, selectmode="day",
            background=PANEL, foreground=FG,
            headersbackground=ENTRY_BG, headersforeground=ACCENT,
            selectbackground=ACCENT, selectforeground=BTN_FG,
            normalbackground=PANEL, normalforeground=FG,
            weekendbackground=PANEL, weekendforeground=FG_DIM,
            othermonthbackground=BG, othermonthforeground=FG_DIM,
            bordercolor=PANEL, font=FONT_LABEL
        )
        self.cal.pack(padx=18, pady=4)

        time_card = tk.Frame(win, bg=PANEL, padx=14, pady=10)
        time_card.pack(fill="x", padx=18, pady=8)

        tk.Label(time_card, text="Select Time", font=FONT_LABEL,
                 bg=PANEL, fg=FG).grid(row=0, column=0, columnspan=6,
                                        sticky="w", pady=(0, 8))

        tk.Label(time_card, text="Hour", font=FONT_SMALL,
                 bg=PANEL, fg=FG_DIM).grid(row=1, column=0, padx=(0, 4))
        self.hour_var = tk.IntVar(value=12)
        tk.Spinbox(time_card, from_=1, to=12, textvariable=self.hour_var,
                   wrap=True, width=4, font=FONT_LABEL,
                   bg=ENTRY_BG, fg=FG, buttonbackground=PANEL,
                   relief="flat").grid(row=1, column=1, padx=(0, 8))

        tk.Label(time_card, text="Min", font=FONT_SMALL,
                 bg=PANEL, fg=FG_DIM).grid(row=1, column=2, padx=(0, 4))
        self.minute_var = tk.IntVar(value=0)
        tk.Spinbox(time_card, from_=0, to=59, textvariable=self.minute_var,
                   wrap=True, width=4, font=FONT_LABEL,
                   bg=ENTRY_BG, fg=FG, buttonbackground=PANEL,
                   relief="flat",
                   format="%02.0f").grid(row=1, column=3, padx=(0, 8))

        self.am_pm_var = tk.StringVar(value="AM")
        am_pm_menu = tk.OptionMenu(time_card, self.am_pm_var, "AM", "PM")
        am_pm_menu.config(
            bg=ENTRY_BG, fg=FG, activebackground=ACCENT,
            activeforeground=BTN_FG, relief="flat",
            font=FONT_LABEL, width=4
        )
        am_pm_menu.grid(row=1, column=4)

        styled_button(win, "✔  Confirm", self.submit_datetime,
                      width=20).pack(pady=(4, 16))

    def submit_datetime(self):
        hour   = self.hour_var.get()
        minute = self.minute_var.get()
        am_pm  = self.am_pm_var.get()

        if am_pm == "PM" and hour != 12:
            hour += 12
        elif am_pm == "AM" and hour == 12:
            hour = 0

        self.selected_time = datetime.time(hour, minute)
        selected_date = self.cal.get_date()   # e.g. "02/24/2026"
        time_str = self.selected_time.strftime("%H:%M")
        full_dt  = f"{selected_date} {time_str}"

        self.datetime_entry.delete(0, tk.END)
        self.datetime_entry.insert(0, full_dt)

        self.cal_window.destroy()
        self._set_status(f"Date & time set: {full_dt}", FG_DIM)


if __name__ == "__main__":
    root = tk.Tk()
    app = XLSXManagerApp(root)
    root.mainloop()
